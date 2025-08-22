#!/usr/bin/env python
import os
import sys
import django
import openpyxl
from datetime import datetime
import re

# Setup Django environment
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'tiketon.settings')
django.setup()

from django.contrib.auth.models import User, Group
from django.db.models import Q
from core.models import Scanner, Event, EventParticipant, TeamLeader

def has_phone_number(text):
    """Check if text contains a phone number pattern"""
    if not text:
        return False
    
    # Check for phone number patterns like +7, 8777, etc.
    return bool(re.search(r'(\+[78]|[78]\d{3}|\d{10})', str(text)))

def extract_name(text):
    """Extract name from text, removing phone numbers"""
    if not text:
        return None
    
    # Convert to string if not already
    text = str(text)
    
    # Remove phone number patterns
    clean_text = re.sub(r'(\+?[78]\d{10}|\d{10}|\+[78]|\d{3}\s\d{3}\s\d{2}\s\d{2})', '', text)
    
    return clean_text.strip()

def find_team_leader(name_text, team_leaders):
    """Find team leader by partial name match"""
    if not name_text:
        return None
    
    print(f"Ищем тимлидера по имени: '{name_text}'")
    name_parts = name_text.lower().split()
    
    # Словарь соответствия русских и английских имен
    name_mapping = {
        'акдаулет': 'akdaulet', 'ospanov': 'оспанов',
        'джамиля': 'dzhamilya', 'turganova': 'турганова',
        'жангерим': 'zhangerim', 'zhusupova': 'жусупова',
        'айгерим': 'aigerim', 'kazbek': 'казбек'
    }
    
    # Расширяем список частей имени, добавляя возможные варианты транслитерации
    expanded_name_parts = []
    for part in name_parts:
        expanded_name_parts.append(part)
        if part in name_mapping:
            expanded_name_parts.append(name_mapping[part])
    
    # Try to find a match among team leaders
    for tl in team_leaders:
        tl_full_name = f"{tl.first_name} {tl.last_name}".lower()
        tl_parts = tl_full_name.split()
        
        # Добавляем возможные варианты транслитерации для имени тимлидера
        expanded_tl_parts = []
        for part in tl_parts:
            expanded_tl_parts.append(part)
            if part in name_mapping:
                expanded_tl_parts.append(name_mapping[part])
        
        # Check if any part of the name matches
        for part in expanded_name_parts:
            if len(part) > 2:  # Avoid matching very short parts
                for tl_part in expanded_tl_parts:
                    if part in tl_part or tl_part in part:
                        print(f"Matched '{part}' with team leader: {tl.first_name} {tl.last_name}")
                        return tl
        
        # Проверка по имени и фамилии отдельно для случаев, когда имя и фамилия поменяны местами
        if any(part in tl.first_name.lower() for part in expanded_name_parts if len(part) > 2) or \
           any(part in tl.last_name.lower() for part in expanded_name_parts if len(part) > 2):
            print(f"Matched by partial name with team leader: {tl.first_name} {tl.last_name}")
            return tl
    
    # Если не нашли по имени, ищем по модели TeamLeader
    for part in expanded_name_parts:
        if len(part) > 2:
            team_leader_model = TeamLeader.objects.filter(
                Q(first_name__icontains=part) | Q(last_name__icontains=part)
            ).first()
            
            if team_leader_model:
                # Ищем соответствующего пользователя
                user = User.objects.filter(
                    Q(first_name__iexact=team_leader_model.first_name) & 
                    Q(last_name__iexact=team_leader_model.last_name)
                ).first()
                
                if user:
                    print(f"Matched via TeamLeader model: {user.first_name} {user.last_name}")
                    return user
    
    print(f"Team leader not found: {name_text}")
    return None

def get_or_create_scanner_for_team_leader(team_leader):
    """
    Находит или создает запись сканера для тимлидера
    """
    # Сначала проверяем, есть ли тимлидер в новой модели TeamLeader
    teamleader_model = TeamLeader.objects.filter(
        first_name=team_leader.first_name,
        last_name=team_leader.last_name
    ).first()
    
    if teamleader_model and teamleader_model.scanner:
        print(f"Найден сканер для тимлидера через модель TeamLeader: {teamleader_model.scanner.first_name} {teamleader_model.scanner.last_name}")
        return teamleader_model.scanner
    
    # Если не найден через модель TeamLeader, ищем напрямую
    scanner = Scanner.objects.filter(
        first_name=team_leader.first_name,
        last_name=team_leader.last_name
    ).first()
    
    # Если сканер не найден, создаем его
    if not scanner:
        scanner = Scanner.objects.create(
            first_name=team_leader.first_name,
            last_name=team_leader.last_name,
            email=team_leader.email
        )
        print(f"Создан сканер для тимлидера: {scanner.first_name} {scanner.last_name}")
        
        # Если есть запись в TeamLeader, связываем с созданным сканером
        if teamleader_model:
            teamleader_model.scanner = scanner
            teamleader_model.save()
            print(f"Сканер связан с тимлидером в модели TeamLeader: {teamleader_model.first_name} {teamleader_model.last_name}")
    else:
        print(f"Найден существующий сканер для тимлидера: {scanner.first_name} {scanner.last_name}")
        
        # Если есть запись в TeamLeader, но нет связи со сканером, создаем её
        if teamleader_model and not teamleader_model.scanner:
            teamleader_model.scanner = scanner
            teamleader_model.save()
            print(f"Сканер связан с существующим тимлидером в модели TeamLeader: {teamleader_model.first_name} {teamleader_model.last_name}")
    
    return scanner

def main():
    print("Starting import of June 2025 events and participants...")
    
    # Path to the Excel file
    excel_path = 'Копия сканеры тикетон астана.xlsx'
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path)
        print(f"Available sheets: {wb.sheetnames}")
        
        # Select the sheet with February 2025 data
        sheet = wb['апрель']  # Use the specific sheet for February 2025
        print(f"Selected sheet: {sheet.title}")
        
        # Get all team leaders from the database
        team_leaders = list(User.objects.filter(groups__name='Тимлидеры'))
        print(f"Found {len(team_leaders)} team leaders in database")
        
        # Print team leaders for debugging
        for tl in team_leaders:
            print(f"DB Team Leader: {tl.first_name} {tl.last_name}")
        
        # Variables to track current event and team leader
        current_event = None
        current_team_leader = None
        
        # Process rows
        for row_idx in range(1, sheet.max_row + 1):
            # Column A: Team leader name and phone
            cell_a = sheet.cell(row=row_idx, column=1).value
            # Column B: Event name
            event_name = sheet.cell(row=row_idx, column=2).value
            # Column C: Date
            event_date_cell = sheet.cell(row=row_idx, column=3).value
            # Column D: Number of scanners
            scanners_count = sheet.cell(row=row_idx, column=4).value
            # Column E: Hours
            event_hours = sheet.cell(row=row_idx, column=5).value
            
            # Check if this row contains a team leader (has phone number)
            if cell_a and has_phone_number(cell_a):
                team_leader_name = extract_name(cell_a)
                if team_leader_name:
                    # Try to find team leader in the database by partial match
                    current_team_leader = find_team_leader(team_leader_name, team_leaders)
            
            # Process event information
            if event_name and event_date_cell and current_team_leader:
                # Convert date to proper format
                if isinstance(event_date_cell, datetime):
                    event_date = event_date_cell.date()
                else:
                    try:
                        # Try to parse date string
                        event_date = datetime.strptime(str(event_date_cell), "%Y-%m-%d").date()
                    except ValueError:
                        try:
                            # Try with day.month.year format
                            event_date = datetime.strptime(str(event_date_cell), "%d.%m.%Y").date()
                        except ValueError:
                            try:
                                # Попытка обработать числовой формат даты Excel
                                if isinstance(event_date_cell, (int, float)):
                                    from datetime import timedelta
                                    # Excel начинает отсчет с 1 января 1900 года
                                    base_date = datetime(1899, 12, 30)  # Excel bug: 1900 не високосный
                                    event_date = (base_date + timedelta(days=int(event_date_cell))).date()
                                else:
                                    print(f"Invalid date format: {event_date_cell}")
                                    continue
                            except Exception as e:
                                print(f"Error parsing date {event_date_cell}: {e}")
                                continue
                
                try:
                    # Create or get the event
                    event, created = Event.objects.get_or_create(
                        name=event_name,
                        date=event_date,
                        created_by=current_team_leader,  # Используем created_by вместо leader
                        defaults={
                            'max_scanners': int(scanners_count) if isinstance(scanners_count, (int, float)) else 10,
                            'duration_hours': float(event_hours) if isinstance(event_hours, (int, float)) else 1.0,
                            'location': 'Астана',  # Значение по умолчанию
                            'description': '',  # Пустое описание
                            'start_date': event_date,  # Устанавливаем start_date равным date
                            'end_date': event_date  # Устанавливаем end_date равным date
                        }
                    )
                    
                    if created:
                        print(f"Created event: {event.name} on {event.date}")
                    else:
                        print(f"Event already exists: {event.name} on {event.date}")
                        # Update fields if needed
                        if isinstance(scanners_count, (int, float)):
                            event.max_scanners = int(scanners_count)
                        if isinstance(event_hours, (int, float)):
                            event.duration_hours = float(event_hours)
                        event.save()
                    
                    current_event = event
                    
                    # Добавляем тимлидера как участника события
                    try:
                        # Получаем или создаем запись сканера для тимлидера
                        team_leader_scanner = get_or_create_scanner_for_team_leader(current_team_leader)
                        
                        # Добавляем тимлидера как участника события
                        team_leader_participant, tl_created = EventParticipant.objects.get_or_create(
                            event=current_event,
                            volunteer=team_leader_scanner
                        )
                        
                        if tl_created:
                            print(f"Тимлидер {current_team_leader.first_name} {current_team_leader.last_name} добавлен как участник события {current_event.name}")
                        
                        # Начисляем часы тимлидеру (можно добавить бонус для тимлидера)
                        team_leader_hours = current_event.duration_hours * 1.5  # Тимлидер получает на 50% больше часов
                        team_leader_participant.hours_awarded = team_leader_hours
                        team_leader_participant.save()
                        print(f"  - Установлено {team_leader_hours} часов для тимлидера {team_leader_scanner.first_name} {team_leader_scanner.last_name}")
                    except Exception as e:
                        print(f"Error adding team leader as participant: {e}")
                except Exception as e:
                    print(f"Error creating/updating event: {e}")
                    continue
            
            # If we have a valid event and the row contains a scanner name (not a team leader row)
            if current_event and cell_a and not has_phone_number(cell_a) and not event_name:
                scanner_name = str(cell_a).strip()
                if scanner_name:
                    try:
                        # Try to find or create scanner
                        names = scanner_name.split()
                        if len(names) >= 1:
                            if len(names) == 1:
                                first_name = names[0]
                                last_name = ""
                            else:
                                first_name = names[0]
                                last_name = ' '.join(names[1:])
                            
                            scanner, created = Scanner.objects.get_or_create(
                                first_name=first_name,
                                last_name=last_name
                            )
                            
                            if created:
                                print(f"Created scanner: {scanner.first_name} {scanner.last_name}")
                            
                            # Add scanner to event if not already added
                            participant, p_created = EventParticipant.objects.get_or_create(
                                event=current_event,
                                volunteer=scanner
                            )
                            
                            if p_created:
                                print(f"Added {scanner.first_name} {scanner.last_name} to {current_event.name}")
                            
                            # Установка часов для участника
                            if current_event.duration_hours:
                                # Учитываем опоздание, если оно есть (предполагаем, что late_minutes может не существовать)
                                late_minutes = 0
                                awarded_hours = current_event.duration_hours
                                
                                # Обновляем часы
                                participant.hours_awarded = awarded_hours
                                participant.save()
                                print(f"  - Установлено {awarded_hours} часов для {scanner.first_name} {scanner.last_name}")
                    except Exception as e:
                        print(f"Error processing scanner {scanner_name}: {e}")
                        continue
        
        # Обновление часов для всех участников ивентов июня 2025
        print("\nОбновление часов для всех участников...")
        # Определяем год и месяц на основе выбранного листа
        if sheet.title.lower() == 'Февраль':
            year = 2024
            month = 2
        elif sheet.title.lower() == 'Март':
            year = 2024
            month = 3
        elif sheet.title.lower() == 'Апрель':
            year = 2024
            month = 4
        elif sheet.title.lower() == 'Май':
            year = 2024
            month = 5
        elif sheet.title.lower() == 'Июнь':
            year = 2024
            month = 6
        elif sheet.title.lower() == 'Июль':
            year = 2024
            month = 7
        elif sheet.title.lower() == 'Август':
            year = 2024
            month = 8
        elif sheet.title.lower() == 'Сентябрь':
            year = 2024
            month = 9
        elif sheet.title.lower() == 'Октябрь':
            year = 2024
            month = 10
        elif sheet.title.lower() == 'Ноябрь':
            year = 2024
            month = 11
        elif sheet.title.lower() == 'Декабрь':
            year = 2024
            month = 12
        elif sheet.title.lower() == 'Январь 25':
            year = 2025
            month = 1
        elif sheet.title.lower() == 'Февраль 25':
            year = 2025
            month = 2
        elif sheet.title.lower() == 'март 25':
            year = 2025
            month = 3
        elif sheet.title.lower() == 'апрель25':
            year = 2025
            month = 4
        elif sheet.title.lower() == 'май25':
            year = 2025
            month = 5
        elif sheet.title.lower() == 'июнь25':
            year = 2025
            month = 6
        else:
            print(f"Не удалось определить год и месяц для листа {sheet.title}")
            year = None
            month = None
        
        if year and month:
            events = Event.objects.filter(date__year=year, date__month=month)
            print(f"Обновление событий за {month}/{year}, найдено событий: {events.count()}")
        else:
            events = Event.objects.all()
            print("Обновление всех событий")
        
        for event in events:
            if event.duration_hours:
                try:
                    participants = EventParticipant.objects.filter(event=event)
                    print(f"Обновление часов для ивента '{event.name}' ({event.date})")
                    print(f"Длительность ивента: {event.duration_hours} часов")
                    print(f"Количество участников: {participants.count()}")
                    
                    # Проверяем, есть ли тимлидер как участник
                    team_leader = event.created_by
                    team_leader_scanner = get_or_create_scanner_for_team_leader(team_leader)
                    
                    # Добавляем тимлидера как участника, если его нет
                    team_leader_participant, tl_created = EventParticipant.objects.get_or_create(
                        event=event,
                        volunteer=team_leader_scanner
                    )
                    
                    if tl_created:
                        print(f"Тимлидер {team_leader.first_name} {team_leader.last_name} добавлен как участник события {event.name}")
                    
                    # Начисляем часы тимлидеру (с бонусом)
                    team_leader_hours = event.duration_hours * 1.5
                    team_leader_participant.hours_awarded = team_leader_hours
                    team_leader_participant.save()
                    print(f"  - Установлено {team_leader_hours} часов для тимлидера {team_leader_scanner.first_name} {team_leader_scanner.last_name}")
                    
                    # Обновляем часы для остальных участников
                    for participant in participants:
                        # Пропускаем тимлидера, так как мы уже обновили его часы
                        if participant.volunteer.id == team_leader_scanner.id:
                            continue
                            
                        # Для обычных сканеров - обычное количество часов
                        awarded_hours = event.duration_hours
                        participant.hours_awarded = awarded_hours
                        participant.save()
                        print(f"  - {participant.volunteer.first_name} {participant.volunteer.last_name}: {awarded_hours} часов")
                except Exception as e:
                    print(f"Error updating hours for event {event.name}: {e}")
                    continue
    
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main() 